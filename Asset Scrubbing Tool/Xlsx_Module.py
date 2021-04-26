#!/usr/bin/env python
# coding: utf-8

# In[ ]:


from openpyxl import Workbook
import os
import openpyxl
import win32com.client as win32
import FileName_Replace
import FileName_Remove


# In[ ]:


def Scrubbing_Process_xlsx(Location,Keywords,Flag,xlsxFileList):       
    
   
    def xlsxElementsRemove(Location,keywords,Flag,xlsxFileList):        
        
        file_path=r"latest_macro.xlsm"
        
        wb = openpyxl.load_workbook(file_path,keep_vba=True)
        
        global FileLocation
        FileLocation = Location
        excel_path = os.path.abspath(Location)

        def add_removewords(sheet_name, column, path):  
            print(column)
            #keys = column.keys()     
            values = column.values()   
            print(values)
            noofreplacement=len(column)             
            ws = wb[sheet_name]    
            new_column = ws.max_column + 1
            ws.cell(row=2,column=2,value=noofreplacement)
            ws.cell(row=1,column=2,value=path)
            #ws.cell(row=3,column=2,value="No")
            for rowy, value in enumerate(column, start=1):
                ws.cell(row=rowy, column=4, value=value)
                new_column = ws.max_column + 1
            for rowy, value in enumerate(values, start=1):
                ws.cell(row=rowy, column=5, value=value)

        
        add_removewords('Macro', keywords, excel_path)
        wb.save(file_path)
        
        def run_excel_macro (file_path, separator_char):
            win32.pythoncom.CoInitialize ()
            xl = win32.Dispatch('Excel.Application')
            xl.Application.visible = False

            try:
                
                wb = xl.Workbooks.Open(os.path.abspath(file_path))
                xl.Application.run(file_path.split(sep=separator_char)[-1] + "!Module1.loopAllSubFolderSelectStartDirectory")
                #print(file_path.split(sep=separator_char)[-1] + "!Module2.KK")
                #xl.Application.run("Trial.xlsm!Module2:KK")
                wb.Save()
                wb.Close()
                print("Run macro")

            except Exception as ex:
                template = "An exception of type {0} occurred. Arguments:\n{1!r}"
                message = template.format(type(ex).__name__, ex.args)
                print(message)

            xl.Application.Quit()
            del xl


        separator_char = os.sep
        
         
#         for i in xlsxFileList:
#             if i.endswith(".xlsx"): 
#                 run_excel_macro(file_path, separator_char)
        run_excel_macro(file_path, separator_char)
    
        def clear_macro(sheet_name):
            ws = wb[sheet_name]
            ws.cell(row=2, column=2).value=None
            ws.cell(row=1, column=2).value=None    
            #ws.cell(row=3,column=2).value=None
            ws.delete_cols(4,5)
    
            #ws.cell(row, column=5).value=None 
        clear_macro('Macro')
        wb.save(file_path)
        
        
    ##Code for Replace Function xlsx
    def xlsxElementsReplace(Location,res,Flag,xlsxFileList):
        file_path=r"latest_macro.xlsm"
        
        keywords = res
        wb = openpyxl.load_workbook(file_path,keep_vba=True)
        
        global FileLocation
        FileLocation= Location
    
        excel_path = os.path.abspath(Location)

        def add_removewords(sheet_name, column, path):    
            keys = column.keys()     
            values = column.values()     
            noofreplacement=len(column)             
            ws = wb[sheet_name]    
            new_column = ws.max_column + 1
            ws.cell(row=2,column=2,value=noofreplacement)
            ws.cell(row=1,column=2,value=path)
            #ws.cell(row=3,column=2,value="No")
            for rowy, value in enumerate(column, start=1):
                ws.cell(row=rowy, column=4, value=value)
                new_column = ws.max_column + 1
            for rowy, value in enumerate(values, start=1):
                ws.cell(row=rowy, column=5, value=value)

        
        add_removewords('Macro', keywords, excel_path)
        wb.save(file_path)
        
        def run_excel_macro (file_path, separator_char):
            win32.pythoncom.CoInitialize ()
            xl = win32.Dispatch('Excel.Application')
            xl.Application.visible = False

            try:
                wb = xl.Workbooks.Open(os.path.abspath(file_path))
                xl.Application.run(file_path.split(sep=separator_char)[-1] + "!Module1.loopAllSubFolderSelectStartDirectory")
                #print(file_path.split(sep=separator_char)[-1] + "!Module2.KK")
                #xl.Application.run("Trial.xlsm!Module2:KK")
                wb.Save()
                wb.Close()

            except Exception as ex:
                template = "An exception of type {0} occurred. Arguments:\n{1!r}"
                message = template.format(type(ex).__name__, ex.args)
                print(message)

            xl.Application.Quit()
            del xl


        separator_char = os.sep

        
#         for i in xlsxFileList:
#             if i.endswith(".xlsx"):                                               
#                 run_excel_macro(file_path, separator_char)
        run_excel_macro(file_path, separator_char)
    
        def clear_macro(sheet_name):
            ws = wb[sheet_name]
            ws.cell(row=2, column=2).value=None
            ws.cell(row=1, column=2).value=None    
            #ws.cell(row=3, column=2).value=None
            ws.delete_cols(4,5)
    
            #ws.cell(row, column=5).value=None 
        clear_macro('Macro')
        wb.save(file_path)
        
    if (Flag=="Remove")|(Flag=="remove")|(Flag=="REMOVE"):
        print("Remove Keywords-")
        Remove_Keywords=Keywords.split(",")
        print(Remove_Keywords)
        
        def Convert(Remove_Keywords):
            res_dct = {Remove_Keywords[i]: "  " for i in range(0, len(Remove_Keywords))}            
            return res_dct              
        keywords = Convert(Remove_Keywords)
        print(keywords)
        
        FileName_Remove.FileName_Removal(Location,Keywords,xlsxFileList)
        xlsxElementsRemove(Location,keywords,Flag,xlsxFileList)
        
    else:
        Replace_List=[]
        print("Replace-")
        Replace_Keywords=Keywords.split(",")
        for i in Replace_Keywords:
            Single_Keyword=i.split(":")
            Replace_List.append(Single_Keyword)
        print(Replace_List)
        
        ini_string1 = Keywords
        res = dict(item.split(":") for item in ini_string1.split(","))         
        
        FileName_Replace.FileName_Replace(Location,Keywords,xlsxFileList)
        xlsxElementsReplace(Location,res,Flag,xlsxFileList)


# In[ ]:




