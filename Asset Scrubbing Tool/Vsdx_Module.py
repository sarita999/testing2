#!/usr/bin/env python
# coding: utf-8

# In[ ]:


from openpyxl import Workbook
import os
import openpyxl
import win32com.client as win32
import FileName_Replace
import FileName_Remove


# In[ ]:


def Scrubbing_Process_vsdx(Location,Keywords,Flag,vsdxFileList):     
    global Scrubbing_Location
    global Scrubbing_Keywords
    global Scrubbing_vsdxFileList
    
    Scrubbing_Location=Location
    Scrubbing_Keywords=Keywords
    Scrubbing_vsdxFileList=vsdxFileList
    
    def vsdxElementsRemove(Location,Keywords,Flag,vsdxFileList):
        print("Yes")
        Remove_Keywords=Scrubbing_Keywords.split(",")
        file_path=r"latest_macro.xlsm"
        Key_List = Keywords.split(",")
        
        def Convert(Key_List):
            res_dct = {Key_List[i]: "  " for i in range(0, len(Key_List))}
            return res_dct
        
        keywords = Convert(Key_List)
        
#         def Convert(Keywords):
#             res_dct = {Keywords[i]: "  " for i in range(0, len(Keywords))}
#             return res_dct
        
#         keywords = Convert(Keywords)
        
        wb = openpyxl.load_workbook(file_path,keep_vba=True)
        
        global FileLocation
        vsdx_path = os.path.abspath(Location)
                    


        def add_removewords(sheet_name, column, path):    
            #keys = column.keys()     
            values = column.values()    
            noofreplacement=len(column)       
           
            ws = wb[sheet_name]    
            new_column = ws.max_column + 1
            ws.cell(row=2,column=2,value=noofreplacement)
            ws.cell(row=1,column=2,value=path)
           # ws.cell(row=3,column=2,value="No")
            for rowy, value in enumerate(column, start=1):
                ws.cell(row=rowy, column=4, value=value)
                new_column = ws.max_column + 1
            for rowy, value in enumerate(values, start=1):
                ws.cell(row=rowy, column=5, value=value)

        add_removewords('Macro', keywords, vsdx_path) 
        wb.save(file_path)

        def run_excel_macro (file_path, separator_char):
            win32.pythoncom.CoInitialize ()
            xl = win32.Dispatch('Excel.Application')
            xl.Application.visible = False

            try:
                wb = xl.Workbooks.Open(os.path.abspath(file_path))
                xl.Application.run(file_path.split(sep=separator_char)[-1] + "!Module1.loopAllSubFolderSelectStartDirectory")
                
                wb.Save()
                wb.Close()

            except Exception as ex:
                template = "An exception of type {0} occurred. Arguments:\n{1!r}"
                message = template.format(type(ex).__name__, ex.args)
                print(message)

            xl.Application.Quit()
            del xl


        separator_char = os.sep
        run_excel_macro(file_path, separator_char)
        
        
#         for i in vsdxFileList:
#             if i.endswith(".vsdx"):     
                                           
#                 run_excel_macro(file_path, separator_char)

        def clear_macro(sheet_name):
            ws = wb[sheet_name]
            ws.cell(row=2, column=2).value=None
            ws.cell(row=1, column=2).value=None    
            #ws.cell(row=3, column=2).value=None
            ws.delete_cols(4,5)
    
            #ws.cell(row, column=5).value=None 
        clear_macro('Macro')
        wb.save(file_path)
        
    def vsdxElementsReplace(Location,res,Flag,vsdxFileList):
        print("Yes")
        Replace_List=[]
        print("Replace-")
        Replace_Keywords=Scrubbing_Keywords.split(",")
        for i in Replace_Keywords:
            Single_Keyword=i.split(":")
            Replace_List.append(Single_Keyword)
        print(Replace_List)
        file_path=r"latest_macro.xlsm"
        
        keywords = res
        wb = openpyxl.load_workbook(file_path,keep_vba=True)
        
        global FileLocation
        vsdx_path = os.path.abspath(Location)
       
        def add_removewords(sheet_name, column, path):    
            keys = column.keys()     
            values = column.values()     
            noofreplacement=len(column)    

            print("yes entered")
            ws = wb[sheet_name]    
            new_column = ws.max_column + 1
            ws.cell(row=2,column=2,value=noofreplacement)
            ws.cell(row=1,column=2,value=path)
            #ws.cell(row=3,column=2,value="No")
            for rowy, value in enumerate(column, start=1):
                ws.cell(row=rowy, column=4, value=value)
            new_column = ws.max_column + 1
            for rowy, value in enumerate(values, start=1):
                ws.cell(row=rowy, column=5, value=value)

        add_removewords('Macro', keywords, vsdx_path)
        wb.save(file_path)

        def run_excel_macro (file_path, separator_char):
            win32.pythoncom.CoInitialize ()
            xl = win32.Dispatch('Excel.Application')
            xl.Application.visible = False

            try:
                wb = xl.Workbooks.Open(os.path.abspath(file_path))
                xl.Application.run(file_path.split(sep=separator_char)[-1] + "!Module1.loopAllSubFolderSelectStartDirectory")
                #print(file_path.split(sep=separator_char)[-1] + "!Module2.KK")
                #xl.Application.run("Trial.xlsm!Module2:KK")
                wb.Save()
                wb.Close()

            except Exception as ex:
                template = "An exception of type {0} occurred. Arguments:\n{1!r}"
                message = template.format(type(ex).__name__, ex.args)
                print(message)

            xl.Application.Quit()
            del xl


        separator_char = os.sep
        run_excel_macro(file_path, separator_char)
        
#         for i in vsdxFileList:
#             if i.endswith(".vsdx"): 
                                             
#                 run_excel_macro(file_path, separator_char)

        def clear_macro(sheet_name):
            ws = wb[sheet_name]
            ws.cell(row=2, column=2).value=None
            ws.cell(row=1, column=2).value=None    
            #ws.cell(row=3, column=2).value=None 
            ws.delete_cols(4,5)

            #ws.cell(row, column=5).value=None 
        clear_macro('Macro')
        wb.save(file_path)
        
        
           
    if (Flag=="Remove")|(Flag=="remove")|(Flag=="REMOVE"):
        print("Remove Keywords-")
        Remove_Keywords=Keywords.split(",")
        print(Remove_Keywords)
        
#         FileName_Remove.FileName_Removal(Location,Keywords,xlsxFileList)
        FileName_Remove.FileName_Removal(Remove_Keywords,Scrubbing_vsdxFileList)
        vsdxElementsRemove(Location,Keywords,Flag,vsdxFileList)
    else:
        Replace_List=[]
        print("Replace-")
        Replace_Keywords=Keywords.split(",")
        for i in Replace_Keywords:
            Single_Keyword=i.split(":")
            Replace_List.append(Single_Keyword)
        print(Replace_List)
        
        ini_string1 = Keywords
        res = dict(item.split(":") for item in ini_string1.split(","))         
        
#         FileName_Replace.FileName_Replace(Location,Keywords,xlsxFileList)
        FileName_Replace.FileName_Replace(Replace_List,Scrubbing_vsdxFileList)
        vsdxElementsReplace(Location,res,Flag,vsdxFileList)

